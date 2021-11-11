from openpyxl import Workbook
from openpyxl import load_workbook

try:
    libro = load_workbook("Mi 1er excel.xlsx")
except FileNotFoundError:
    print("El archivo no existe, adiós!")
    exit()
    
print("Cantidad de hojas:", len(libro.sheetnames))
numHoja = 0
for nombreHoja in libro.sheetnames:
    print("Hoja",numHoja+1,nombreHoja)
    if nombreHoja == "Alumnos":
        libro.active = numHoja
    numHoja += 1
hoja = libro.active
print ("La página activa es:",hoja.title)
#Una forma de leer
rango = hoja["A2":"C4"]#Adaptar rango a totalidad de celdas con info
for celda in rango:
    for objeto in celda:
        print(objeto,"contiene",objeto.value)
input()
#Otra opción
for fila in rango:
    for columna in fila:
        print(columna,"contiene:",columna.value)
input()
#Leer por filas
print ("Impresión de todas las celdas")
for fila in hoja.values:
    for valor in fila:
        print (valor,end= "\t")
    print ("\n")
