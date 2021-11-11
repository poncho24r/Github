from pyhunter import PyHunter
from openpyxl import Workbook
import getpass


def Busqueda(organizacion):
    #Cantidad de resultados esperados de la búsqueda
    #El límite MENSUAL de Hunter es 50, cuidado!
    resultado = hunter.domain_search (company = organizacion, limit = 1, emails_type = 'personal')
    return resultado

def GuardarInformacion(datosEncontrados,organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    #Agrega el codigo necesario para guardar en formato tabla
    #dentro del libro de Excel, información que consideres relevante
    #de lo obtenido en la búsqueda.
    a1 = hoja['A1']
    a1.value = "DOMINIO"
    b1 = hoja['B1']
    b1.value = "ORGANIZACION"
    c1 = hoja['C1']
    c1.value = "PAIS"
    
    hoja.cell(2,1,datosEncontrados["domain"])
    hoja.cell(2,2,datosEncontrados["organization"])
    hoja.cell(2,3,datosEncontrados["country"])
    
    libro.save("Hunter" + organizacion + ".xlsx")

print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter (apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)
if datosEncontrados == None:
    exit()
else:
    print(datosEncontrados)
    print(type(datosEncontrados))
    GuardarInformacion(datosEncontrados,orga)



