from DicAlumnos import IPGpo09
from openpyxl import Workbook
from openpyxl import load_workbook

try:
    libro = load_workbook("Mi 1er excel.xlsx")
except FileNotFoundError:
    print("El archivo no existe, adiós!")
    exit()

if len(libro.sheetnames)>1:
    libro.active = 1
    hoja = libro.active
else:
    hoja = libro.active
print ("La página activa es:",hoja.title)

celda = hoja["A1"]
hoja["A1"] = "Num. lista"
print ("Contenido de A1:",celda.value)
hoja["B1"] = "Matricula"
hoja["C1"] = "Alumno"
print ("Los encabezados de las celdas son:")
print (hoja["A1"].value,end = "\t")
print (hoja["B1"].value, end = "\t")
print (hoja["C1"].value)

count = 2
for x,y in IPGpo09.items():
    hoja.cell(count,1,x)
    hoja.cell(count,2,y[0])
    hoja.cell(count,3,y[1])
    count += 1

libro.save("Mi 1er excel.xlsx")
