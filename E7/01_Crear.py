#!/usr/bin/env python3.7
from openpyxl import Workbook #parcial
#import openpyxl  #total

#se "abre" excel y se tiene un libro nuevo y en blanco
libro = Workbook() #libro = openxyl.Workbook()
pagina = libro.active
pagina1 = libro.create_sheet("Pagina1")
cadena = input("Nombre de la pagina: ")
pagina1.title = cadena
print(libro.sheetnames)
libro.save("Mi 1er excel.xlsx")

